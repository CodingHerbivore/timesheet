import random
import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook, Workbook

# Change the workbook to whatever you intend to use as your 'input'
# TODO: Change this to select from an 'input' folder
book = load_workbook('example-input.xlsx')
sheet = book.active
names = []
tasks = []
starting_time = datetime.datetime(2021, 7, 1, 0, 0, 0)

workday_start = 8  # Enter the hour that you start the workday. Must be in 24-hour format.
workday_end = 18  # Enter the hour that you end the day. Must be in 24-hour format.
work_day = starting_time + datetime.timedelta(hours=workday_start)

# Make a list of all the names in our input timesheet
# TODO: Ignore headers
for cell in sheet['B']:
    if cell.value:
        if not cell.value in names:
            names.append(cell.value)

# Make a list of all the tasks in our import timesheet
# TODO: Ignore headers
for cell in sheet['C']:
    if cell.value:
        if not cell.value in tasks:
            tasks.append(cell.value)

# Create the new timesheet
timebook = Workbook()
timesheet = timebook.active

# Make column headers
timesheet["A1"] = "Day"
timesheet["B1"] = "Customer"
timesheet["C1"] = "Task"
timesheet["D1"] = "Time Started"
timesheet["E1"] = "Time Ended"

# Start data at row 2
timesheetrow = 2


# Determine how long the task will take depending on the remaining time in the day
def task_length(remainder):
    if remainder > 180:
        find_duration: int = random.randint(0, 2)

        if find_duration == 0:
            minutes = random.randint(5, 30)
        elif find_duration == 1:
            minutes = random.randint(5, 60)
        elif find_duration == 2:
            minutes = random.randint(5, 180)
    else:
        minutes = random.randint(5, remainder)

    # Make sure we're working in 5-minute increments
    while minutes % 5 != 0:
        minutes += 1

    return minutes


# Change "months=1" to "days=5" (or 7) if you use weekly timesheets and vice-versa
while not work_day > starting_time + relativedelta(days=5):

    def day(workday_start, workday_end, working_time):
        workday_end = working_time + datetime.timedelta(hours=(workday_end - workday_start))

        while (workday_end - working_time).total_seconds() / 60 != 0:
            global timesheetrow
            remainder = (workday_end - working_time).total_seconds() / 60

            day = working_time
            timesheet["A" + str(timesheetrow)] = day.date()
            print('Task Start ' + str(working_time))
            begin = working_time.time()
            timesheet["D" + str(timesheetrow)] = "{:d}:{:02d}".format(begin.hour, begin.minute)

            task_minutes = task_length(remainder)

            working_time += datetime.timedelta(minutes=task_minutes)
            print('Task End ' + str(working_time))
            end = working_time.time()
            timesheet["E" + str(timesheetrow)] = "{:d}:{:02d}".format(end.hour, end.minute)

            task_customer = str(random.choices(names))[2:-2]

            timesheet["B" + str(timesheetrow)] = task_customer
            print('Customer ' + task_customer)

            task = str(random.choices(tasks))[2:-2]

            print('Task ' + task)
            timesheet["C" + str(timesheetrow)] = task
            print('-----')
            timesheetrow += 1


    day(workday_start, workday_end, work_day)
    work_day = work_day + datetime.timedelta(days=1)
timebook.save("generated-timesheet.xlsx")
